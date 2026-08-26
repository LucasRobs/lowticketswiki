#!/usr/bin/env python3
"""
build_outputs.py - Gera planilha, CSV, favoritos e painel a partir do achados.json
Uso: python3 build_outputs.py achados.json --saida ./radar --historico ./Radar/radar-low-ticket.xlsx
"""

import json
import sys
import argparse
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_historico(historico_path):
    """Carrega o histórico anterior da planilha."""
    if not os.path.exists(historico_path):
        return {}
    wb = load_workbook(historico_path)
    ws = wb['Achados'] if 'Achados' in wb.sheetnames else wb.active
    
    historico = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value:
            termo = str(row[0].value)  # termo_busca
            primeira_vez = row[8].value if len(row) > 8 and row[8].value else None
            rodadas = row[9].value if len(row) > 9 and row[9].value else 0
            historico[termo] = {
                'primeira_vez_visto': primeira_vez,
                'rodadas_visto': rodadas
            }
    return historico

def load_historico_aba(historico_path):
    """Carrega o histórico da aba 'Historico'."""
    if not os.path.exists(historico_path):
        return {}
    wb = load_workbook(historico_path)
    if 'Historico' not in wb.sheetnames:
        return {}
    ws = wb['Historico']
    
    historico = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            historico[str(row[0])] = {
                'data_primeira': row[1],
                'rodadas': row[2]
            }
    return historico

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('achados_json')
    parser.add_argument('--saida', default='./radar')
    parser.add_argument('--historico', default=None)
    parser.add_argument('--painel', action='store_true')
    args = parser.parse_args()
    
    with open(args.achados_json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    data_varredura = data.get('data_varredura', datetime.now().strftime('%Y-%m-%d'))
    achados = data.get('achados', [])
    
    # Carrega histórico
    historico_primeira = {}
    historico_rodadas = {}
    if args.historico and os.path.exists(args.historico):
        h = load_historico(args.historico)
        for k, v in h.items():
            historico_primeira[k] = v.get('primeira_vez_visto')
            historico_rodadas[k] = v.get('rodadas_visto', 0)
        
        h2 = load_historico_aba(args.historico)
        for k, v in h2.items():
            if k not in historico_primeira:
                historico_primeira[k] = v.get('data_primeira')
            if k not in historico_rodadas:
                historico_rodadas[k] = v.get('rodadas', 0)
    
    # Prepara saída
    os.makedirs(args.saida, exist_ok=True)
    
    # === PLANILHA XLSX ===
    xlsx_path = os.path.join(args.saida, f'radar-low-ticket.xlsx')
    
    if os.path.exists(args.historico if args.historico else ''):
        wb = load_workbook(args.historico)
        # Remove aba Achados antiga se existir
        if 'Achados' in wb.sheetnames:
            del wb['Achados']
        if 'Resumo_Nicho' in wb.sheetnames:
            del wb['Resumo_Nicho']
        if 'Historico' in wb.sheetnames:
            del wb['Historico']
    else:
        wb = Workbook()
        # Remove sheet padrão
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    
    # --- Aba Achados ---
    ws = wb.create_sheet('Achados')
    
    headers = ['termo_busca', 'produto', 'tipo', 'gateway', 'nicho', 'temperatura', 
               'mencoes', 'faixa_preco', 'primeira_vez_visto', 'rodadas_visto', 
               'descricao', 'evidencia_url']
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Dados
    for row_idx, achado in enumerate(achados, 2):
        termo = achado.get('termo_busca', '')
        primeira = historico_primeira.get(termo, data_varredura)
        rodadas = historico_rodadas.get(termo, 0) + 1
        
        values = [
            termo,
            achado.get('produto', ''),
            achado.get('tipo', ''),
            achado.get('gateway', ''),
            achado.get('nicho', ''),
            achado.get('temperatura', ''),
            achado.get('mencoes', 0),
            achado.get('faixa_preco', ''),
            primeira,
            rodadas,
            achado.get('descricao', ''),
            achado.get('evidencia_url', '')
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajusta largura das colunas
    col_widths = [20, 30, 10, 15, 25, 12, 10, 20, 15, 12, 60, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(achados)+1}"
    
    # --- Aba Resumo por Nicho ---
    ws2 = wb.create_sheet('Resumo_Nicho')
    nicho_stats = {}
    for achado in achados:
        nicho = achado.get('nicho', 'Sem nicho')
        temp = achado.get('temperatura', 'fria')
        if nicho not in nicho_stats:
            nicho_stats[nicho] = {'quente': 0, 'morna': 0, 'fria': 0, 'total': 0, 'marca': 0, 'angulo': 0}
        nicho_stats[nicho][temp] += 1
        nicho_stats[nicho]['total'] += 1
        if achado.get('tipo') == 'marca':
            nicho_stats[nicho]['marca'] += 1
        else:
            nicho_stats[nicho]['angulo'] += 1
    
    headers2 = ['nicho', 'total', 'marca', 'angulo', 'quente', 'morna', 'fria']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, (nicho, stats) in enumerate(sorted(nicho_stats.items(), key=lambda x: -x[1]['total']), 2):
        ws2.cell(row=row_idx, column=1, value=nicho).border = thin_border
        ws2.cell(row=row_idx, column=2, value=stats['total']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=stats['marca']).border = thin_border
        ws2.cell(row=row_idx, column=4, value=stats['angulo']).border = thin_border
        ws2.cell(row=row_idx, column=5, value=stats['quente']).border = thin_border
        ws2.cell(row=row_idx, column=6, value=stats['morna']).border = thin_border
        ws2.cell(row=row_idx, column=7, value=stats['fria']).border = thin_border
    
    for i in range(1, 8):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    
    # --- Aba Historico ---
    ws3 = wb.create_sheet('Historico')
    headers3 = ['termo_busca', 'data_primeira', 'rodadas']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Combina histórico antigo com novo
    all_terms = set(list(historico_primeira.keys()) + [a.get('termo_busca', '') for a in achados])
    for row_idx, termo in enumerate(sorted(all_terms), 2):
        ws3.cell(row=row_idx, column=1, value=termo).border = thin_border
        ws3.cell(row=row_idx, column=2, value=historico_primeira.get(termo, data_varredura)).border = thin_border
        ws3.cell(row=row_idx, column=3, value=historico_rodadas.get(termo, 0)).border = thin_border
    
    for i in range(1, 4):
        ws3.column_dimensions[get_column_letter(i)].width = 25
    
    wb.save(xlsx_path)
    print(f"Planilha salva: {xlsx_path}")
    
    # === CSV ===
    csv_path = os.path.join(args.saida, f'radar-low-ticket-{data_varredura}.csv')
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write(','.join(headers) + '\n')
        for achado in achados:
            termo = achado.get('termo_busca', '')
            primeira = historico_primeira.get(termo, data_varredura)
            rodadas = historico_rodadas.get(termo, 0) + 1
            row = [
                termo,
                achado.get('produto', '').replace(',', ';'),
                achado.get('tipo', ''),
                achado.get('gateway', ''),
                achado.get('nicho', '').replace(',', ';'),
                achado.get('temperatura', ''),
                str(achado.get('mencoes', 0)),
                achado.get('faixa_preco', '').replace(',', ';'),
                primeira,
                str(rodadas),
                achado.get('descricao', '').replace(',', ';').replace('\n', ' '),
                achado.get('evidencia_url', '')
            ]
            f.write(','.join(f'"{v}"' for v in row) + '\n')
    print(f"CSV salvo: {csv_path}")
    
    # === FAVORITOS HTML ===
    fav_path = os.path.join(args.saida, f'favoritos-radar-{data_varredura}.html')
    
    # Agrupa por nicho
    by_nicho = {}
    for achado in achados:
        nicho = achado.get('nicho', 'Sem nicho')
        if nicho not in by_nicho:
            by_nicho[nicho] = []
        by_nicho[nicho].append(achado)
    
    html = '''<!DOCTYPE NETSCAPE-Bookmark-file-1>
<META HTTP-EQUIV="Content-Type" CONTENT="text/html; charset=UTF-8">
<TITLE>Bookmarks</TITLE>
<H1>Bookmarks</H1>
<DL><p>
'''
    
    for nicho, items in sorted(by_nicho.items()):
        html += f'    <DT><H3 ADD_DATE="{int(datetime.now().timestamp())}" LAST_MODIFIED="{int(datetime.now().timestamp())}">{nicho}</H3>\n'
        html += '    <DL><p>\n'
        for achado in items:
            url = achado.get('evidencia_url', '')
            title = f"[{achado.get('temperatura', '').upper()}] {achado.get('produto', '')} ({achado.get('gateway', '')})"
            if achado.get('faixa_preco'):
                title += f" - {achado.get('faixa_preco')}"
            add_date = int(datetime.now().timestamp())
            html += f'        <DT><A HREF="{url}" ADD_DATE="{add_date}" TAGS="radar,lowticket,{achado.get("gateway","").lower()},{achado.get("tipo","")}">{title}</A>\n'
        html += '    </DL><p>\n'
    
    html += '</DL><p>\n'
    
    with open(fav_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"Favoritos salvos: {fav_path}")
    
    # === PAINEL HTML ===
    if args.painel:
        panel_path = os.path.join(args.saida, f'radar-{data_varredura}.html')
        
        # Contadores
        total = len(achados)
        quente = sum(1 for a in achados if a.get('temperatura') == 'quente')
        morna = sum(1 for a in achados if a.get('temperatura') == 'morna')
        fria = sum(1 for a in achados if a.get('temperatura') == 'fria')
        marcas = sum(1 for a in achados if a.get('tipo') == 'marca')
        angulos = sum(1 for a in achados if a.get('tipo') == 'angulo')
        
        panel_html = f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Radar Low Ticket - {data_varredura}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0d1117; color: #e6edf3; line-height: 1.6; padding: 20px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        header {{ margin-bottom: 30px; border-bottom: 1px solid #30363d; padding-bottom: 20px; }}
        h1 {{ font-size: 2rem; font-weight: 600; }}
        .meta {{ color: #8b949e; margin-top: 8px; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin-bottom: 30px; }}
        .stat-card {{ background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 20px; }}
        .stat-value {{ font-size: 2.5rem; font-weight: 700; color: #58a6ff; }}
        .stat-label {{ color: #8b949e; font-size: 0.9rem; margin-top: 4px; }}
        .stat-card.quente .stat-value {{ color: #f85149; }}
        .stat-card.morna .stat-value {{ color: #d29922; }}
        .stat-card.fria .stat-value {{ color: #3fb950; }}
        table {{ width: 100%; border-collapse: collapse; background: #161b22; border-radius: 8px; overflow: hidden; }}
        th {{ background: #21262d; color: #e6edf3; padding: 12px 16px; text-align: left; font-weight: 600; border-bottom: 2px solid #30363d; }}
        td {{ padding: 12px 16px; border-bottom: 1px solid #21262d; }}
        tr:hover {{ background: #1f2428; }}
        .badge {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; }}
        .badge-quente {{ background: #f85149; color: white; }}
        .badge-morna {{ background: #d29922; color: black; }}
        .badge-fria {{ background: #3fb950; color: white; }}
        .badge-marca {{ background: #58a6ff; color: white; }}
        .badge-angulo {{ background: #a371f7; color: white; }}
        .gateway-pp {{ color: #58a6ff; }}
        .gateway-ck {{ color: #f78166; }}
        a {{ color: #58a6ff; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .desc {{ color: #8b949e; font-size: 0.85rem; margin-top: 4px; }}
        .price {{ font-family: 'SF Mono', monospace; color: #7ee787; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📡 Radar Low Ticket</h1>
            <div class="meta">Varredura: {data_varredura} | Gateways: PerfectPay, Cakto | {total} achados ({marcas} marcas, {angulos} ângulos)</div>
        </header>
        
        <div class="stats">
            <div class="stat-card quente">
                <div class="stat-value">{quente}</div>
                <div class="stat-label">Quente (4+ menções)</div>
            </div>
            <div class="stat-card morna">
                <div class="stat-value">{morna}</div>
                <div class="stat-label">Morna (2-3 menções)</div>
            </div>
            <div class="stat-card fria">
                <div class="stat-value">{fria}</div>
                <div class="stat-label">Fria (1 menção)</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Produto</th>
                    <th>Tipo</th>
                    <th>Gateway</th>
                    <th>Nicho</th>
                    <th>Temp.</th>
                    <th>Menções</th>
                    <th>Faixa Preço</th>
                    <th>Termo Busca</th>
                    <th>Descrição</th>
                    <th>Evidência</th>
                </tr>
            </thead>
            <tbody>
'''
        
        for achado in achados:
            temp = achado.get('temperatura', 'fria')
            tipo = achado.get('tipo', '')
            gateway = achado.get('gateway', '')
            gateway_class = 'gateway-pp' if gateway == 'PerfectPay' else 'gateway-ck'
            
            panel_html += f'''                <tr>
                    <td><strong>{achado.get('produto', '')}</strong></td>
                    <td><span class="badge badge-{tipo}">{tipo}</span></td>
                    <td class="{gateway_class}">{gateway}</td>
                    <td>{achado.get('nicho', '')}</td>
                    <td><span class="badge badge-{temp}">{temp}</span></td>
                    <td>{achado.get('mencoes', 0)}</td>
                    <td class="price">{achado.get('faixa_preco', '')}</td>
                    <td><code>{achado.get('termo_busca', '')}</code></td>
                    <td><div class="desc">{achado.get('descricao', '')}</div></td>
                    <td><a href="{achado.get('evidencia_url', '')}" target="_blank">🔗 RA</a></td>
                </tr>
'''
        
        panel_html += '''            </tbody>
        </table>
        
        <div style="margin-top: 30px; padding: 20px; background: #161b22; border: 1px solid #30363d; border-radius: 8px; color: #8b949e; font-size: 0.9rem;">
            <strong>⚠️ Lembrete:</strong> Gateway não é vendedor. PerfectPay, Cakto e afins processam pagamento de produtos de terceiros; aparecer aqui não diz nada sobre a idoneidade delas.
            <br><br>
            <strong>Amostra:</strong> 1 página de cada gateway (5 reclamações por página, ~1 dia de cobertura). A lista do Reclame Aqui não gira em 24h — rodar diariamente produz a mesma fotografia.
        </div>
    </div>
</body>
</html>
'''
        
        with open(panel_path, 'w', encoding='utf-8') as f:
            f.write(panel_html)
        print(f"Painel salvo: {panel_path}")

if __name__ == '__main__':
    main()